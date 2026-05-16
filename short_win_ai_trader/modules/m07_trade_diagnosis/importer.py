"""交割单导入解析器 — 支持Excel/CSV/手动"""

import csv
import os
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

from ...core.logger import get_logger
from ...data_platform.data_models import TradeRecord

logger = get_logger("swat.trade_import")


class TradeImporter:
    """交易记录导入器
    
    支持主流券商导出格式:
    - Excel (.xlsx): 同花顺/东方财富/通达信等
    - CSV (.csv): 通用格式
    - 手动录入: 单笔输入
    """

    # 常见列名映射
    COLUMN_MAP = {
        "代码": "ticker",
        "股票代码": "ticker",
        "证券代码": "ticker",
        "名称": "stock_name",
        "股票名称": "stock_name",
        "证券名称": "stock_name",
        "日期": "trade_date",
        "成交日期": "trade_date",
        "时间": "trade_time",
        "成交时间": "trade_time",
        "操作": "trade_type",
        "买卖": "trade_type",
        "成交类别": "trade_type",
        "价格": "price",
        "成交价格": "price",
        "成交价": "price",
        "数量": "volume",
        "成交数量": "volume",
        "成交量": "volume",
        "金额": "amount",
        "成交金额": "amount",
        "成交额": "amount",
        "盈亏": "profit_loss",
        "实现盈亏": "profit_loss",
        "模式": "trade_mode",
        "交易方式": "trade_mode",
    }

    TRADE_TYPE_MAP = {
        "买入": "买入",
        "买": "买入",
        "证券买入": "买入",
        "卖出": "卖出",
        "卖": "卖出",
        "证券卖出": "卖出",
        "buy": "买入",
        "sell": "卖出",
    }

    def __init__(self):
        logger.info("交割单导入器初始化完成")

    def import_from_file(self, file_path: str) -> List[TradeRecord]:
        """从文件导入交割单"""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        ext = path.suffix.lower()
        if ext == ".xlsx":
            return self._import_excel(file_path)
        elif ext == ".csv":
            return self._import_csv(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {ext}，请使用.xlsx或.csv")

    def _import_excel(self, file_path: str) -> List[TradeRecord]:
        """导入Excel格式"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("请安装openpyxl: pip install openpyxl")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1]]
        column_mapping = self._map_columns(headers)

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = self._parse_row(row, column_mapping)
            if record:
                records.append(record)

        logger.info(f"Excel导入完成: {len(records)}条记录")
        return records

    def _import_csv(self, file_path: str) -> List[TradeRecord]:
        """导入CSV格式"""
        records = []
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)
            column_mapping = self._map_columns(headers)

            for row in reader:
                record = self._parse_row(row, column_mapping)
                if record:
                    records.append(record)

        logger.info(f"CSV导入完成: {len(records)}条记录")
        return records

    def _map_columns(self, headers: List[str]) -> Dict[int, str]:
        """映射列名到标准字段"""
        mapping = {}
        for i, header in enumerate(headers):
            if header in self.COLUMN_MAP:
                mapping[i] = self.COLUMN_MAP[header]
        return mapping

    def _parse_row(self, row: tuple, column_mapping: Dict[int, str]) -> Optional[TradeRecord]:
        """解析单行数据"""
        data = {}
        for col_idx, field_name in column_mapping.items():
            if col_idx < len(row):
                data[field_name] = row[col_idx]

        if not data.get("ticker"):
            return None

        # 处理ticker格式
        ticker = str(data.get("ticker", "")).strip()
        if "." not in ticker and len(ticker) == 6:
            # 自动补充后缀
            if ticker.startswith("6"):
                ticker = f"{ticker}.SH"
            elif ticker.startswith("0") or ticker.startswith("3"):
                ticker = f"{ticker}.SZ"
            elif ticker.startswith("8") or ticker.startswith("4"):
                ticker = f"{ticker}.BJ"

        # 解析日期
        trade_date = data.get("trade_date")
        if trade_date:
            if isinstance(trade_date, str):
                for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%m/%d/%Y"]:
                    try:
                        trade_date = datetime.strptime(trade_date, fmt)
                        break
                    except ValueError:
                        continue

        # 解析交易类型
        trade_type = self.TRADE_TYPE_MAP.get(
            str(data.get("trade_type", "")).strip(),
            str(data.get("trade_type", "买入")).strip(),
        )

        # 解析价格
        try:
            price = float(data.get("price", 0))
        except (ValueError, TypeError):
            price = 0.0

        # 解析数量
        try:
            volume = int(float(str(data.get("volume", "0")).replace(",", "")))
        except (ValueError, TypeError):
            volume = 0

        # 解析金额
        try:
            amount = float(str(data.get("amount", "0")).replace(",", ""))
        except (ValueError, TypeError):
            amount = price * volume

        # 解析盈亏
        profit_loss = None
        if "profit_loss" in data and data["profit_loss"]:
            try:
                profit_loss = float(str(data["profit_loss"]).replace(",", ""))
            except (ValueError, TypeError):
                pass

        # 盈亏比例
        profit_loss_pct = None
        if profit_loss and amount > 0:
            profit_loss_pct = round(profit_loss / amount * 100, 2)

        # 交易模式自动识别
        trade_mode = data.get("trade_mode", "")
        if not trade_mode:
            trade_mode = self._auto_detect_mode(price, volume, amount)

        return TradeRecord(
            trade_id=str(uuid.uuid4())[:8],
            ticker=ticker,
            stock_name=str(data.get("stock_name", "")),
            trade_date=trade_date if isinstance(trade_date, datetime) else datetime.now(),
            trade_type=trade_type,
            price=price,
            volume=volume,
            amount=amount,
            trade_mode=trade_mode,
            profit_loss=profit_loss,
            profit_loss_pct=profit_loss_pct,
        )

    def _auto_detect_mode(self, price: float, volume: int, amount: float) -> str:
        """自动识别交易模式"""
        # 简单启发式判断
        if amount > 1000000:  # 100万以上
            return "打板"
        elif volume > 10000:  # 大手数
            return "半路"
        return "低吸"

    def manual_entry(
        self,
        ticker: str,
        stock_name: str,
        trade_date: str,
        trade_type: str,
        price: float,
        volume: int,
        trade_mode: str = "",
        profit_loss: Optional[float] = None,
    ) -> TradeRecord:
        """手动录入单笔交易"""
        if "." not in ticker and len(ticker) == 6:
            if ticker.startswith("6"):
                ticker = f"{ticker}.SH"
            elif ticker.startswith("0") or ticker.startswith("3"):
                ticker = f"{ticker}.SZ"
            elif ticker.startswith("8") or ticker.startswith("4"):
                ticker = f"{ticker}.BJ"

        dt = datetime.strptime(trade_date, "%Y-%m-%d") if trade_date else datetime.now()
        amount = price * volume

        profit_loss_pct = None
        if profit_loss:
            profit_loss_pct = round(profit_loss / amount * 100, 2)

        if not trade_mode:
            trade_mode = self._auto_detect_mode(price, volume, amount)

        return TradeRecord(
            trade_id=str(uuid.uuid4())[:8],
            ticker=ticker,
            stock_name=stock_name,
            trade_date=dt,
            trade_type=trade_type,
            price=price,
            volume=volume,
            amount=amount,
            trade_mode=trade_mode,
            profit_loss=profit_loss,
            profit_loss_pct=profit_loss_pct,
        )
